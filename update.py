#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Token 出海数据库 — 网页自动更新脚本

工作流程：
  1. 扫描 ../Tokens/ 找到最新的加密 Excel（按文件名日期）
  2. 用密码 ainb 解密
  3. 从「看板」sheet 抽取手动写的文字（标题/总结/KPI/模块/前10排名）
  4. 从两个数据 sheet 抽取历史序列（总量曲线、国产模型明细、份额）
  5. 生成 data.json 供网页读取
  6. （可选）git 提交并推送到 GitHub Pages

用法：
  python3 update.py              # 仅生成 data.json（本地兜底数据）
  python3 update.py --firebase   # 生成并写入 Firebase（看板实时刷新，推荐）
  python3 update.py --push       # 生成并推送到 GitHub Pages（旧工作流）
"""

import sys
import os
import re
import io
import json
import glob
import datetime
from pathlib import Path

# ---------- 路径定位（基于脚本自身位置，不依赖工作目录）----------
SCRIPT_DIR = Path(__file__).resolve().parent
TOKENS_DIR = SCRIPT_DIR.parent / "Tokens"
DATA_JSON  = SCRIPT_DIR / "data.json"

EXCEL_PASSWORD = "ainb"

# 国产厂商配置（与 js/models-config.js 保持一致；新增厂商时两边都加）
# name: 显示名；aliases: Excel 表头可能用的别名（小写匹配）
MODELS_CONFIG = [
    {"name": "DeepSeek", "aliases": ["deepseek", "ds"]},
    {"name": "Qwen",     "aliases": ["qwen", "通义", "通义千问"]},
    {"name": "Kimi",     "aliases": ["kimi", "moonshot"]},
    {"name": "GLM",      "aliases": ["glm", "智谱", "zhipu", "chatglm"]},
    {"name": "Minimax",  "aliases": ["minimax", "mini-max"]},
    {"name": "小米",      "aliases": ["小米", "mimo", "mi-mo", "redmi"]},
    {"name": "腾讯",      "aliases": ["腾讯", "hy", "hunyuan", "混元"]},
]
# 默认列映射（仅当表头扫描失败时回退使用，1-based，M=13 是日期列）
DEFAULT_MODEL_COLS = [
    ("DeepSeek", 14, 15), ("Qwen", 16, 17), ("Kimi", 18, 19),
    ("GLM", 20, 21), ("Minimax", 22, 23), ("小米", 24, 25), ("腾讯", 26, 27),
]

# ---------- 依赖检查 ----------
def ensure_deps():
    missing = []
    try:
        import openpyxl  # noqa
    except ImportError:
        missing.append("openpyxl")
    try:
        import msoffcrypto  # noqa
    except ImportError:
        missing.append("msoffcrypto-tool")
    if missing:
        print("缺少依赖：" + "、".join(missing))
        print("请运行：pip3 install " + " ".join(missing))
        sys.exit(1)


# ---------- 1. 找最新 Excel ----------
def find_latest_excel():
    if not TOKENS_DIR.exists():
        print(f"找不到 Tokens 文件夹：{TOKENS_DIR}")
        sys.exit(1)
    candidates = []
    for f in TOKENS_DIR.glob("*.xlsx"):
        m = re.search(r"(\d{8})", f.name)
        if m:
            candidates.append((m.group(1), f))
    if not candidates:
        print(f"在 {TOKENS_DIR} 中没有找到带日期的 .xlsx 文件")
        sys.exit(1)
    candidates.sort(key=lambda x: x[0], reverse=True)
    return candidates[0]  # (datestr, path)


# ---------- 2. 解密 ----------
def decrypt_to_workbook(xlsx_path):
    import msoffcrypto
    import openpyxl
    tmp = io.BytesIO()
    with open(xlsx_path, "rb") as f:
        office = msoffcrypto.OfficeFile(f)
        office.load_key(password=EXCEL_PASSWORD)
        office.decrypt(tmp)
    tmp.seek(0)
    return openpyxl.load_workbook(tmp, data_only=True)


# ---------- 工具函数 ----------
def col_to_idx(letter):
    """A->1, Z->26, AA->27"""
    n = 0
    for c in letter.upper():
        n = n * 26 + (ord(c) - ord("A") + 1)
    return n

def cell_text(ws, coord):
    v = ws[coord].value
    return "" if v is None else str(v)

def cell_lines(ws, coord):
    """读取单元格文本，按换行拆成非空行列表"""
    t = cell_text(ws, coord)
    return [ln.strip() for ln in re.split(r"[\r\n]+", t) if ln.strip()]

def fmt_date(v):
    if v is None:
        return None
    if isinstance(v, (datetime.datetime, datetime.date)):
        return v.strftime("%Y-%m-%d")
    return str(v)

def to_num(v):
    """转 float，失败返回 None"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace(",", "").replace("%", "")
    try:
        return float(s)
    except ValueError:
        return None

def normalize_wow(v):
    """把环比规整为百分比数值（10.0 表示 +10%）。无法解析返回 None"""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        x = float(v)
        return round(x * 100, 2) if abs(x) <= 1 else round(x, 2)
    s = str(v).strip().replace(",", "")
    m = re.search(r"([+-]?\d+\.?\d*)\s*%?", s)
    if not m:
        return None
    return round(float(m.group(1)), 2)

def parse_wow_text(text):
    """从 '环比-4%' / '+86%' 这类文本提取 (数值, 是否上升)"""
    if not text:
        return (None, None)
    m = re.search(r"([+-]?\d+\.?\d*)", text)
    if not m:
        return (None, None)
    num = float(m.group(1))
    up = num >= 0
    return (round(num, 2), up)

def match_model_by_header(text):
    """用 Excel 表头文字匹配厂商，返回显示名；匹配不到返回 None。
       表头不区分大小写、去空格（与 js/models-config.js 的 matchModelByHeader 一致）。"""
    if not text:
        return None
    h = str(text).strip().lower()
    for m in MODELS_CONFIG:
        cands = [m["name"].lower()] + [a.lower() for a in m.get("aliases", [])]
        if h in cands:
            return m["name"]
    return None


# ---------- 3. 抽取「看板」sheet 文字 ----------
def extract_dashboard_text(wb):
    ws = wb["看板"]

    title = cell_text(ws, "W2").strip()
    summary_raw = cell_text(ws, "W3")
    summary = [ln.strip() for ln in re.split(r"[\r\n]+", summary_raw) if ln.strip()]

    # KPI 卡片
    kpi_raw = [cell_text(ws, "W6"), cell_text(ws, "AB6"), cell_text(ws, "AG6")]
    kpi = []
    for raw in kpi_raw:
        lines = [ln.strip() for ln in re.split(r"[\r\n]+", raw) if ln.strip()]
        label = lines[0] if len(lines) > 0 else ""
        value = lines[1] if len(lines) > 1 else ""
        trend = lines[2] if len(lines) > 2 else ""
        kpi.append({"label": label, "value": value, "trend": trend})

    # 模块标题（标题行可能含两行：标题 + 一句话点评）
    mod_coords = ["W11", "W41", "W55", "W117"]
    modules = []
    for c in mod_coords:
        lines = [ln.strip() for ln in re.split(r"[\r\n]+", cell_text(ws, c)) if ln.strip()]
        modules.append({
            "title": lines[0] if lines else "",
            "comment": lines[1] if len(lines) > 1 else ""
        })

    # 前 10 排名卡片（按位置确定排名 1-10）
    # 单元格文本结构固定为 4 行：第X名 / 模型名 / N.NT Tokens / 环比X%
    rank_coords = [
        "W45", "Z45", "AC45", "AF45", "AI45",   # 1-5
        "W50", "Z50", "AC50", "AF50", "AI50",   # 6-10
    ]
    top10 = []
    for i, c in enumerate(rank_coords):
        lines = [ln.strip() for ln in re.split(r"[\r\n]+", cell_text(ws, c)) if ln.strip()]
        rank = i + 1
        # 默认按行顺序取值；若某行缺失则向后顺延容错
        name, tokens_text, wow_text = "", "", ""
        non_rank_lines = [ln for ln in lines if not re.match(r"^第[一二三四五六七八九十百\d]+名$", ln)]
        if len(non_rank_lines) >= 1:
            name = non_rank_lines[0]
        if len(non_rank_lines) >= 2:
            tokens_text = non_rank_lines[1]
        if len(non_rank_lines) >= 3:
            wow_text = non_rank_lines[2]
        wow_num, up = parse_wow_text(wow_text)
        # 从 '4.70T Tokens' 提取数字
        tnum = None
        mt = re.search(r"(\d+\.?\d*)\s*T", tokens_text)
        if mt:
            tnum = float(mt.group(1))
        top10.append({
            "rank": rank,
            "name": name or f"第{rank}名",
            "tokens_text": tokens_text,
            "tokens": tnum,
            "wow_text": wow_text,
            "wow_num": wow_num,
            "up": up,
        })

    footer_raw = cell_text(ws, "W179")
    footer = re.sub(r"\s{2,}", "  ", footer_raw).strip()

    return {
        "title": title,
        "summary": summary,
        "kpi": kpi,
        "modules": modules,
        "top10": top10,
        "footer": footer,
    }


# ---------- 4. 抽取历史数据 ----------
def extract_total_history(wb):
    """【原始数据】消耗量统计：周总量曲线"""
    name = "【原始数据】消耗量统计"
    if name not in wb.sheetnames:
        return {"dates": [], "values": [], "wow": []}
    ws = wb[name]
    dates, values, wow = [], [], []
    row = 2
    while True:
        d = ws.cell(row=row, column=1).value
        v = ws.cell(row=row, column=2).value
        w = ws.cell(row=row, column=3).value
        if d is None and v is None:
            break
        dates.append(fmt_date(d))
        values.append(to_num(v))
        wow.append(normalize_wow(w))
        row += 1
        if row > 500:
            break
    return {"dates": dates, "values": values, "wow": wow}


def extract_domestic(wb):
    """【国内】每天消耗量统计 的周聚合块 (M:AE, 行 2-60)
       原始单位为「十亿 token」，÷1000 转 T，与总量曲线保持一致。

       列映射优先扫描第 1 行表头动态识别（新增厂商只要表头能匹配别名即可自动生效），
       扫描不到任何模型时回退到 DEFAULT_MODEL_COLS。"""
    name = "【国内】每天消耗量统计"
    if name not in wb.sheetnames:
        return {}
    ws = wb[name]

    # ---- 扫描表头（第 1 行），动态识别列映射 ----
    model_cols = {}                      # {model_name: (值列, 环比列)}
    total_vcol, total_wcol = 28, 29      # 默认 AB 国产总和 / AC 环比
    global_vcol = 30                     # 默认 AD 全球总和
    share_col = 31                       # 默认 AE 份额
    max_col = ws.max_column or 0
    for c in range(14, max_col + 1):
        h = ws.cell(row=1, column=c).value
        if h is None:
            continue
        m = match_model_by_header(h)
        if m and m not in model_cols:
            # 模型值列 = 命中列，环比列 = 紧邻的下一列
            model_cols[m] = (c, c + 1)
            continue
        text = str(h)
        if "国产" in text and "总和" in text:
            total_vcol = c
            total_wcol = c + 1
        elif "全球" in text:
            global_vcol = c
        elif "份额" in text:
            share_col = c
    # 至少匹配到一个模型才算扫描成功，否则整体回退默认列
    if not model_cols:
        model_cols = {m: (v, w) for (m, v, w) in DEFAULT_MODEL_COLS}

    # 所有配置厂商都出现在输出里（即使该 Excel 无此厂商数据，也按行填 None）
    models = {m["name"]: {"values_T": [], "wow": []} for m in MODELS_CONFIG}
    dates = []
    total_T, total_wow, global_T, share = [], [], [], []

    row = 2
    while True:
        d = ws.cell(row=row, column=13).value  # M
        if d is None:
            # 再多看一行，避免空行干扰
            nxt = ws.cell(row=row + 1, column=13).value
            if nxt is None:
                break
            row += 1
            continue
        dates.append(fmt_date(d))
        # 每一行都为全部配置厂商填一格：命中的取值，未命中填 None，保证各模型数组等长
        for m in models:
            if m in model_cols:
                vcol, wcol = model_cols[m]
                raw = to_num(ws.cell(row=row, column=vcol).value)
                models[m]["values_T"].append(round(raw / 1000, 3) if raw is not None else None)
                models[m]["wow"].append(normalize_wow(ws.cell(row=row, column=wcol).value))
            else:
                models[m]["values_T"].append(None)
                models[m]["wow"].append(None)
        tot = to_num(ws.cell(row=row, column=total_vcol).value)   # 国产总和
        total_T.append(round(tot / 1000, 3) if tot is not None else None)
        total_wow.append(normalize_wow(ws.cell(row=row, column=total_wcol).value))
        g = to_num(ws.cell(row=row, column=global_vcol).value)    # 全球总和
        global_T.append(round(g / 1000, 3) if g is not None else None)
        sh = to_num(ws.cell(row=row, column=share_col).value)     # 份额
        share.append(round(sh * 100, 2) if (sh is not None and abs(sh) <= 1) else (sh if sh is not None else None))
        row += 1
        if row > 200:
            break

    return {
        "dates": dates,
        "models": models,
        "total_T": total_T,
        "total_wow": total_wow,
        "global_T": global_T,
        "share": share,
    }


# ---------- 5. 组装 + 写出 ----------
def build_json(wb, source_file, week_label):
    dash = extract_dashboard_text(wb)
    data = {
        "meta": {
            "source_file": source_file,
            "week_label": week_label,
            "generated_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        },
        **dash,
        "total_history": extract_total_history(wb),
        "domestic": extract_domestic(wb),
    }
    return data

def write_json(data):
    with open(DATA_JSON, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    return DATA_JSON


# ---------- 6. Git 推送 ----------
def git_push(week_label):
    import subprocess
    def run(cmd):
        print(f"  $ {' '.join(cmd)}")
        return subprocess.run(cmd, cwd=str(SCRIPT_DIR), capture_output=True, text=True)
    # 确保 repo 已初始化
    if not (SCRIPT_DIR / ".git").exists():
        print("⚠️  尚未初始化 Git 仓库，跳过推送。请先按 README 完成 GitHub 首次部署。")
        return False
    # 检查是否已配置远程仓库
    remotes = run(["git", "remote"])
    if not (r_stdout := (remotes.stdout or "").strip()):
        print("⚠️  尚未配置 GitHub 远程仓库，跳过推送。")
        print("   请先按 README.md「首次部署」把项目推到 GitHub，之后每周双击即可自动更新。")
        return False
    run(["git", "add", "data.json"])
    run(["git", "commit", "-m", f"更新数据 {week_label}"])
    r = run(["git", "push"])
    if r.returncode == 0:
        print("✅ 已推送到 GitHub，稍等片刻网页自动更新。")
        return True
    else:
        print("⚠️  推送失败：" + (r.stderr or r.stdout))
        return False


# ---------- 6b. Firebase 写入（REST API，无需额外依赖）----------
FIREBASE_CONFIG_PATH = SCRIPT_DIR / "js" / "firebase-config.js"

def _read_firebase_config():
    """从 js/firebase-config.js 解析出 databaseURL；返回 None 表示未配置。"""
    if not FIREBASE_CONFIG_PATH.exists():
        return None
    txt = FIREBASE_CONFIG_PATH.read_text(encoding="utf-8")
    m = re.search(r'databaseURL\s*:\s*"([^"]+)"', txt)
    return m.group(1) if m else None

def push_to_firebase(data):
    """用 Firebase REST API 覆盖写入数据。
       需要 Realtime Database 处于「测试模式」（30 天内允许读写）或配置了规则。"""
    import urllib.request
    db_url = _read_firebase_config()
    if not db_url or "YOUR_" in db_url:
        print("⚠️  Firebase 未配置（js/firebase-config.js 里仍是占位符）。")
        print("   请先按 README.md 完成 Firebase 配置。本次只生成 data.json，不写云端。")
        return False
    # 拆成 content / charts / settings 三块写入
    content = {
        "title": data.get("title"),
        "summary": data.get("summary"),
        "kpi": data.get("kpi"),
        "modules": data.get("modules"),
        "top10": data.get("top10"),
        "footer": data.get("footer"),
        "meta": data.get("meta"),
    }
    charts = {
        "total_history": data.get("total_history"),
        "domestic": data.get("domestic"),
    }
    settings = {"updated_at": datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "admin_user": "admin"}
    base = db_url.rstrip("/") + "/"
    for key, payload in [("content", content), ("charts", charts), ("settings", settings)]:
        body = json.dumps(payload, ensure_ascii=False).encode("utf-8")
        req = urllib.request.Request(base + key + ".json",
                                     data=body,
                                     headers={"Content-Type": "application/json"},
                                     method="PUT")
        try:
            with urllib.request.urlopen(req, timeout=20) as resp:
                resp.read()
            print(f"   ✓ {key} 已写入")
        except Exception as e:
            print(f"   ✗ {key} 写入失败：{e}")
            print(f"     提示：若为权限错误，请在 Firebase 控制台把 Realtime Database 规则临时设为公开测试模式。")
            return False
    print("✅ 已写入 Firebase，看板将实时刷新。")
    return True


# ---------- 主流程 ----------
def main():
    do_push = "--push" in sys.argv
    do_firebase = "--firebase" in sys.argv
    ensure_deps()

    print("🔍 正在查找最新数据库...")
    week_label, xlsx_path = find_latest_excel()
    print(f"   找到：{xlsx_path.name}（周期 {week_label}）")

    print("🔓 解密中...")
    wb = decrypt_to_workbook(xlsx_path)
    print(f"   工作表：{wb.sheetnames}")

    print("📊 提取数据...")
    data = build_json(wb, xlsx_path.name, week_label)

    # 简要校验
    n_total = len(data["total_history"]["values"])
    n_dom = len(data["domestic"].get("dates", []))
    print(f"   看板：标题/KPI/{len(data['top10'])}个排名/4个模块")
    print(f"   总量历史：{n_total} 周")
    print(f"   国产明细：{n_dom} 周 × {len(data['domestic'].get('models', {}))} 个模型")

    # 总是先生成 data.json（作为兜底数据源 + 备份）
    print("💾 生成 data.json...")
    out = write_json(data)
    print(f"   已写入：{out}")

    # --firebase：写入 Firebase 云数据库（看板实时刷新）
    if do_firebase:
        print("☁️  写入 Firebase...")
        push_to_firebase(data)

    # --push：git 推送（旧 GitHub Pages 工作流，兼容保留）
    if do_push and not do_firebase:
        print("🚀 推送到 GitHub Pages...")
        git_push(week_label)

    print("🎉 完成！")
    return data


if __name__ == "__main__":
    main()
